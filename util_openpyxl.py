# Version 6/18/24 Add autofit_column_width
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import openpyxl.utils as pyxl_util

def open_wb(sfile):
    """
    open workbook and return openpyxl wb object
    JDL 3/16/23
    """
    return openpyxl.load_workbook(sfile)

def delete_sht(wb, sht):
    """
    Delete specified sheet from a workbook
    JDL 3/16/23
    """
    if sht in wb.sheetnames: wb.remove(wb[sht])
    return wb
    
def write_df_as_wb_sht(sfile, sht, df, is_index=False):
    """
    write DataFrame to specified sheet in Excel Workbook
    (Delete previous version of sheet first to write as replacement)
    JDL 3/16/23
    """
    with pd.ExcelWriter(sfile, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=sht, index=is_index)

def clear_worksheet(ws):
    """
    Clear Excel worksheet (ws object in an openpyxl wb object)
    JDL 4/21/23
    """
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    return ws

def clear_cell(cell):
    """
    openyxl clear a cell's value, formatting and cell comment/note
    """
    cell.value = None
    cell.font = openpyxl.styles.Font()
    cell.border = openpyxl.styles.Border()
    cell.fill = openpyxl.styles.PatternFill()
    cell.number_format = 'General'
    cell.alignment = openpyxl.styles.Alignment()
    if cell.comment: cell.comment = None
    return cell

def ws_to_df(ws):
    """
    Convert an openpyxl ws to a DataFrame (range index and columns)
    """
    data = ws.values
    df = pd.DataFrame(data)
    return df

def clear_columns(ws, col1, col2):
    """
    Clear specified columns
    """
    for col in range(col1, col2+1):
        for row in range(1, ws.max_row + 1):
            clear_cell(ws.cell(row=row, column=col))
    return ws

def find_string_in_row(ws, irow, sfind):
    """
    Find cell with specified string in specified row
    JDL 4/25/23
    """
    for c in ws[irow]:
        if c.value == sfind: return c
    return None

def find_string_in_col(ws, icol, sfind):
    """
    Find cell with specified string in specified column
    JDL 4/25/23
    """
    for col in ws.iter_cols(min_col=icol, max_col=icol):
        for c in col: 
            if c.value == sfind: return c
    return None

def write_lst_to_rng(ws, cell_home, lstvals, direction='row'):
    """
    Write list of values to cells in specified row or column on openpyxl ws
    direction: either 'row' or 'col'
    JDL 4/25/23
    """
    for i, val in enumerate(lstvals):
        if direction == 'row':
            ws.cell(row=cell_home.row, column=cell_home.column+i, value=val)
        elif direction == 'col':
            ws.cell(row=cell_home.row+i, column=cell_home.column, value=val)
    return ws

def toggle_sheet_visibility(wb, sht, IsHide=True):
    """
    Toggle the visibility of a sheet in an openpyxl workbook.
    JDL 5/4/23
    wb: the openpyxl Workbook object
    sht: [String] name of sheet to toggle visibility
    IsHide: [Boolean] toggle to hide or unhide (default is hide)
    """
    ws = wb[sht]
    if IsHide:
        ws.sheet_state = 'hidden'
    else:
        ws.sheet_state = 'visible'
""" 
===============================================================================
Range iterators
===============================================================================
"""

def rng_iterator(ws, cell_home, cell_end):
    """
    Return row-wise iterator to iterate over cells in range
    specified by openpyxl home and end cells. Usage: for c in cell_iterator(xxx):
    JDL 4/21/23
    """
    row_start, col_start = cell_home.row, cell_home.column
    row_end, col_end = cell_end.row, cell_end.column

    for row in range(row_start, row_end+1):
        for col in range(col_start, col_end+1):
            cell = ws.cell(row=row, column=col)
            yield cell
            
def rng_iterator_enum(ws, cell_home, cell_end):
    """
    Return row-wise iterator with row, column enumeration to iterate 
    over cells in a range specified by openpyxl home and end cells.
    Usage: for i, j, c in cell_iterator(xxx): where i and j are the
    row and column indices of cells c returned by the generator
    JDL 4/21/23
    """
    start_row, start_col = cell_home.row, cell_home.column
    end_row, end_col,  = cell_end.row, cell_end.column, 
    for i, row in enumerate(range(start_row, end_row+1), start=1):
        for j, col in enumerate(range(start_col, end_col+1), start=1):
            cell = ws.cell(row=row, column=col)
            yield (i, j, cell)

""" 
===============================================================================
Functions for writing DataFrame values and setting dict for df cell locations
===============================================================================
"""

def write_dataframe(ws, df, cell_home):
    """
    Write a DataFrame to a specific openpyxl cell on an Excel ws
    cell_home argument is ws.cell for top left data cell in Excel
    JDL 4/23/23
    """
    #Create dict of cell locations for df elements
    d_cells = set_df_openpyxl_cell_locns(ws, df, cell_home)
    
    #Write data, index and column values
    for fn in [write_df_data, write_df_index, write_df_columns]:
        ws = fn(ws, df, d_cells)
    return ws 

def set_df_openpyxl_cell_locns(ws, df, cell_home):
    """
    Set ws.cells for ranges of data, index and columns
    cell_home argument is ws.cell for top left data cell in Excel
    JDL 4/23/23; Add n_col_lvls 3/11/24
    """
    row, col = row_col(cell_home)
    d_cells = {'cell_home_data':cell_home}
    d_cells['cell_end_data'] = ws.cell(row + df.index.size - 1, col + df.columns.size - 1)
    d_cells['cell_home_idx'] = ws.cell(row, col - 1)
    d_cells['cell_end_idx'] = ws.cell(row + df.index.size - 1, col - 1)
    d_cells['cell_home_cols'] = ws.cell(row - 1, col)
    d_cells['cell_end_cols'] = ws.cell(row - 1, col + df.columns.size - 1)

    #Added 3/11/24 for multiindex cols - should refactor to just used cell_home_cols
    d_cells['n_col_lvls'] = set_num_col_levels(df)
    d_cells['cell_cols_rng_begin'] = set_cols_rng_begin(ws, d_cells)
    d_cells['cell_cols_rng_end'] = d_cells['cell_end_cols']
    return d_cells

def set_num_col_levels(df):
    """
    return number of column levels for single or multiindex columns
    JDL 3/11/24
    """
    return len(df.columns.levels) if isinstance(df.columns, pd.MultiIndex) else 1

def offset_cell(ws, cell, irows=0, icols=0):
    """
    Return a cell that is irows below the specified cell
    JDL 3/11/24
    """
    return ws.cell(row=cell.row + irows, column=cell.column + icols)

def set_cols_rng_begin(ws, d_cells):
    """
    Return column values end cell for single or multiindex columns
    JDL 3/11/24
    """
    if d_cells['n_col_lvls'] == 0 :
        return d_cells['cell_home_cols']
    else:
        return offset_cell(ws, d_cells['cell_home_cols'], irows= - d_cells['n_col_lvls'] + 1)

def set_cols_rng_end(ws, d_cells):
    """
    Return column values end cell for single or multiindex columns
    JDL 3/11/24
    """
    if d_cells['n_col_lvls'] == 0 :
        return d_cells['cell_end_cols']
    else:
        return offset_cell(ws, d_cells['cell_end_cols'], irows=-1)

def row_col(c):
    """
    return openpyxl ws.cell row and column tuple
    JDL 4/23/23
    """
    return c.row, c.column 
    
def write_df_data(ws, df, d_cells):
    """
    Write DataFrame's data values
    JDL 4/23/23
    """
    for i, j, c in rng_iterator_enum(ws, d_cells['cell_home_data'], d_cells['cell_end_data']):
        c.value = df.values[i-1][j-1]        
    return ws
    
def write_df_index(ws, df, d_cells):
    """
    Write DataFrame's index to column adjacent to first data column
    JDL 4/23/23
    """
    #Write index values
    for i, j, c in rng_iterator_enum(ws, d_cells['cell_home_idx'], d_cells['cell_end_idx']):
        c.value = list(df.index)[i-1]
    
    #Write index name as heading above index values
    ws.cell(d_cells['cell_home_idx'].row - 1, d_cells['cell_home_idx'].column).value = df.index.name
    return ws

def write_df_columns(ws, df, d_cells):
    """
    Write DataFrame's column values to row above to first data row
    JDL 4/23/23
    """    
    for i, j, c in rng_iterator_enum(ws, d_cells['cell_home_cols'], d_cells['cell_end_cols']):
        c.value = list(df.columns)[j-1]  
    return ws

""" 
===============================================================================
Functions for writing DataFrames with multiindex columns
Temp - refactor into basic functions for writing dataframes
===============================================================================
"""
def write_dataframe_multi_cols(ws, df, cell_home):
    """
    Write a DataFrame with single or multiindex cols to a specific openpyxl 
    cell on an Excel ws cell_home argument is ws.cell for top left data cell in Excel
    (temporary - refactor write_dataframe for this to be base approach)
    JDL 3/11/23
    """
    #Create dict of cell locations for df elements
    d_cells = set_df_openpyxl_cell_locns(ws, df, cell_home)
    
    #Write data, index and column values
    ws = write_df_data(ws, df, d_cells)
    ws = write_df_index(ws, df, d_cells)
    ws = write_df_columns_multi(ws, df, d_cells, IsMergeMatching=True)
    return ws 

def write_df_columns_multi(ws, df, d_cells, IsMergeMatching=False):
    """
    Write DataFrame's column values to row above to first data row (works for multiindex cols)
    JDL 3/11/24
    """    
    n_lvls = d_cells['n_col_lvls']
    for i, j, c in rng_iterator_enum(ws, d_cells['cell_home_cols'], d_cells['cell_end_cols']):
        if n_lvls == 1:
            c.value = list(df.columns)[j-1]
        else:
            #column label tuple
            column = df.columns[j-1]
            for lvl in range(1, n_lvls+1): 
                offset_cell(ws, c, -lvl+1).value = column[n_lvls - lvl]

                #n_lvls = 3
                #lvl = 1, offset(0), column[2] --> offset(-lvl + 1), column(n_lvls - lvl)
                #lvl = 2, offset(-1), column[1]
                #lvl = 3, offset(-2), column[0]

                #n_lvls = 2
                #lvl = 1, offset(0), column[1] --> offset(-lvl + 1), column(n_lvls - lvl)
                #lvl = 2, offset(-1), column[1]
    if IsMergeMatching: merge_matching(ws, d_cells)
    return ws

def merge_matching(ws, d_cells):
    """
    Merge cells with matching, adjacent multiindex column labels
    JDL 3/11/24
    """

    #Set local names for column limits
    col_home, col_end = d_cells['cell_home_cols'].column, d_cells['cell_end_cols'].column

    for lvl in range(1, d_cells['n_col_lvls']+1):
        
        #Set row number based on columns home row and iteration counter
        row = d_cells['cell_home_cols'].row - lvl + 1

        #Initialize block of values to check and iterate
        prev_value, start_col = None, 1
        for col in range(col_home, col_end + 1):
            cell_value = ws.cell(row, col).value

            #If current cell value is same as the previous, continue (potential merge)
            if cell_value == prev_value:
                continue

            #check for merge (if more than one cell) and re-initialize
            else:
                if col - start_col > 1: merge_cell_sequence(ws, row, start_col, row, col-1)
                start_col, prev_value = col, cell_value

        #Check for sequence at the end of range
        if col_end - start_col > 0: merge_cell_sequence(ws, row, start_col, row, col_end)

def merge_cell_sequence(ws, row1, col1, row2, col2):
    """
    Merge specified range
    JDL 3/11/24
    """
    ws.merge_cells(start_row=row1, start_column=col1, end_row=row2, end_column=col2)

"""
Formatting multiindex columns
"""
def set_df_borders_multi(ws, df, cell_home):
    """
    Set borders for an Excel range containing a DataFrame
    (temp version to also handle multiindex columns)
    JDL 3/11/24
    """
    d_cells = set_df_openpyxl_cell_locns(ws, df, cell_home)
    ws = set_df_data_borders(ws, d_cells, 'thin')
    ws = set_df_index_borders(ws, d_cells, 'thin')
    ws = set_df_cols_borders_multi(ws, d_cells, 'thin')
    return ws

def set_df_cols_borders_multi(ws, d_cells, style_border):
    """
    """
    cell_rng_begin = d_cells['cell_cols_rng_begin']
    set_range_border(ws, cell_rng_begin, d_cells['cell_end_cols'], style_border)
    return ws

def set_df_builtin_styles_multi(ws, df, cell_home, style_data=None, style_idx=None, style_cols=None):
    """
    Set built-in styles for Excel range with a DataFrame
    (temp version to also handle multiindex columns)
    JDL 3/11/24
    """
    d_cells = set_df_openpyxl_cell_locns(ws, df, cell_home)
    if not style_data is None: ws = set_df_data_builtin_styles(ws, d_cells, style_data)
    if not style_idx is None: ws = set_df_index_builtin_styles(ws, d_cells, style_idx)
    if not style_cols is None: ws = set_df_cols_builtin_styles_multi(ws, d_cells, style_cols)
    return ws

def set_df_cols_builtin_styles_multi(ws, d_cells, style_cols):
    """
    Set built-in Excel style for df column values and index name cell
    JDL 3/11/24 - should refactor to consolidate with set_df_cols_builtin_styles
    """
    cell_rng_begin = d_cells['cell_cols_rng_begin']
    set_range_builtin_style(ws, cell_rng_begin, d_cells['cell_end_cols'], style_cols)
    
    #Set index name cell same style as data columns
    row_start = d_cells['cell_home_idx'].row - d_cells['n_col_lvls']
    row_end = d_cells['cell_home_idx'].row - 1
    col = d_cells['cell_home_idx'].column
    set_range_builtin_style(ws, ws.cell(row_start, col), ws.cell(row_end, col), style_cols)
    return ws

def set_df_cols_align_multi(ws, d_cells, d_align):
    """
    Set alignment df column values
    """
    row_start = d_cells['cell_home_idx'].row - d_cells['n_col_lvls']
    row_end = d_cells['cell_home_idx'].row - 1
    cell_start = ws.cell(row_start, d_cells['cell_home_cols'].column)
    cell_end = ws.cell(row_end, d_cells['cell_end_cols'].column)

    set_range_alignment(ws, cell_start, cell_end, d_align)
    return ws

def autofit_column_width(ws, col):
    """
    Autofit the width of a column in an openpyxl worksheet.
    col is column number
    JDL 6/18/24   
    """
    col_letter = pyxl_util.get_column_letter(col)
    column = ws[col_letter]

    #Get the max length of any cell value in the column
    max_length = 0
    for cell in column:
        try:
            if len(str(cell.value)) > max_length: max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[col_letter].width = adjusted_width
""" 
===============================================================================
Functions for setting borders
===============================================================================
"""
def set_openpyxl_border_obj(style_border):
    """
    Create a border style based on style_border='thick', 'thin' etc.
    Use "from openpyxl.styles import Border, Side" to import needed openpyxl attributes
    JDL 4/21/23
    """
    return Border(left=Side(style=style_border),
                  right=Side(style=style_border), 
                  top=Side(style=style_border),
                  bottom=Side(style=style_border))

def set_range_border(ws, cell_home, cell_end, style_border):
    """
    Set borders for an Excel range defined by ws cell_home and cell_end
    JDL 4/21/23
    """
    #Create a Border object for style_border
    border_obj = set_openpyxl_border_obj(style_border)
    
    #Apply the border_obj to each cell in the range
    for c in rng_iterator(ws, cell_home, cell_end):
        c.border = border_obj
        
def set_df_borders(ws, df, cell_home):
    """
    Set borders for an Excel range containing a DataFrame
    JDL 4/21/23
    """
    d_cells = set_df_openpyxl_cell_locns(ws, df, cell_home)
    ws = set_df_data_borders(ws, d_cells, 'thin')
    ws = set_df_index_borders(ws, d_cells, 'thin')
    ws = set_df_cols_borders(ws, d_cells, 'thick')
    return ws

def set_df_data_borders(ws, d_cells, style_border):
    """
    Put border around cells for df data values
    """
    set_range_border(ws, d_cells['cell_home_data'], d_cells['cell_end_data'], style_border)
    return ws

def set_df_index_borders(ws, d_cells, style_border):
    set_range_border(ws, d_cells['cell_home_idx'], d_cells['cell_end_idx'], style_border)
    row = d_cells['cell_home_idx'].row - 1
    col = d_cells['cell_home_idx'].column
    set_range_border(ws, ws.cell(row, col), ws.cell(row, col), style_border)
    return ws

def set_df_cols_borders(ws, d_cells, style_border):
    set_range_border(ws, d_cells['cell_home_cols'], d_cells['cell_end_cols'], style_border)
    return ws

""" 
===============================================================================
Functions for setting built-in styles for a range or a DataFrame
===============================================================================
"""

def set_range_builtin_style(ws, cell_home, cell_end, style_builtin):
    """
    Apply the builtin style to each cell in the range
    JDL 4/25/23
    """
    for c in rng_iterator(ws, cell_home, cell_end):
        c.style = style_builtin
        
def set_df_builtin_styles(ws, df, cell_home, style_data=None, style_idx=None, style_cols=None):
    """
    Set built-in styles for Excel range with a DataFrame
    JDL 4/25/23
    """
    d_cells = set_df_openpyxl_cell_locns(ws, df, cell_home)
    if not style_data is None: ws = set_df_data_builtin_styles(ws, d_cells, style_data)
    if not style_idx is None: ws = set_df_index_builtin_styles(ws, d_cells, style_idx)
    if not style_cols is None: ws = set_df_cols_builtin_styles(ws, d_cells, style_cols)
    return ws

def set_df_data_builtin_styles(ws, d_cells, style_data):
    """
    Set built-in Excel style for df data values
    """
    set_range_builtin_style(ws, d_cells['cell_home_data'], d_cells['cell_end_data'], style_data)
    return ws

def set_df_index_builtin_styles(ws, d_cells, style_idx):
    """
    Set built-in Excel style for df index values
    """    
    set_range_builtin_style(ws, d_cells['cell_home_idx'], d_cells['cell_end_idx'], style_idx)
    return ws

def set_df_cols_builtin_styles(ws, d_cells, style_cols, fmt_idx_name=True):
    """
    Set built-in Excel style for df column values and index name cell
    """    
    set_range_builtin_style(ws, d_cells['cell_home_cols'], d_cells['cell_end_cols'], style_cols)
    
    #Set index name cell same style as data columns
    if fmt_idx_name:
        row = d_cells['cell_home_idx'].row - 1
        col = d_cells['cell_home_idx'].column
        set_range_builtin_style(ws, ws.cell(row, col), ws.cell(row, col), style_cols)
    return ws

""" 
===============================================================================
Functions for setting column widths in an openpyxl ws object
===============================================================================
"""

def set_range_column_widths(ws, col_start, col_end, width):
    """
    Set a contiguous range of columns (e.g. df.columns) to a specified width
    JDL 4/25/23
    """
    for col in range(col_start, col_end+1):
        letter = pyxl_util.get_column_letter(col)
        ws.column_dimensions[letter].width = width + 0.6
    return ws

"""
===============================================================================
Functions for setting cell alignment properties
===============================================================================
"""

def set_range_alignment(ws, cell_home, cell_end, d_align):
    """
    Set alignment for an Excel range defined by ws cell_home and cell_end
    JDL 6/29/23
    """
    for c in rng_iterator(ws, cell_home, cell_end):
        align = Alignment()
        if 'horizontal' in d_align: align.horizontal = d_align['horizontal']
        if 'wrap_text' in d_align: align.wrap_text = d_align['wrap_text']
        c.alignment = align                                        
        
def set_df_data_align(ws, d_cells, d_align):
    """
    Set alignment df data values
    """
    set_range_alignment(ws, d_cells['cell_home_data'], d_cells['cell_end_data'], d_align)
    return ws

def set_df_index_align(ws, d_cells, d_align):
    """
    Set alignment df index values and name
    """
    set_range_alignment(ws, d_cells['cell_home_idx'], d_cells['cell_end_idx'], d_align)
    row = d_cells['cell_home_idx'].row - 1
    col = d_cells['cell_home_idx'].column
    set_range_alignment(ws, ws.cell(row, col), ws.cell(row, col), d_align)
    return ws

def set_df_cols_align(ws, d_cells, d_align):
    """
    Set alignment df column values
    """
    set_range_alignment(ws, d_cells['cell_home_cols'], d_cells['cell_end_cols'], d_align)
    return ws
""" 
===============================================================================
Functions for setting number formats in openpyxl ws object
===============================================================================
"""
def set_range_num_format(ws, cell_home, cell_end, 
                         num_fmt='General', num_fmt_zeros='General'):
    """
    Apply Excel number format to each cell in a range
    JDL 12/5/23
    """
    for c in rng_iterator(ws, cell_home, cell_end):
        c.number_format = num_fmt
        if c.value == 0: c.number_format = num_fmt_zeros

def set_cell_num_format(cell, num_format):
    """
    Set Excel cell number format
    JDL 12/5/23
    """
    cell.number_format = num_format